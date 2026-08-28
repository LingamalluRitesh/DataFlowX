"""
DataFlowX Microsoft Excel Connector
Supports multi-sheet workbook inspection, streaming row reading via openpyxl, and chunking.
"""

from datetime import datetime
import os
import time
from typing import Any, Dict, Generator, List, Optional
import openpyxl
import pandas as pd
from backend.core.exceptions import ConnectorError
from backend.core.logging import get_logger
from connectors.base import (
    BaseConnector,
    ColumnMeta,
    ConnectionTestResult,
    ExtractionChunk,
    FieldType,
    SchemaDiscoveryResult,
    TableMeta,
)

logger = get_logger(__name__)


class ExcelConnector(BaseConnector):
    """Production connector for Excel (.xlsx / .xls) files."""

    def __init__(self, config: Dict[str, Any], credentials: Optional[Dict[str, Any]] = None):
        super().__init__(config, credentials)
        self.connector_type = "excel"
        self.file_path = self.config.get("file_path", "")
        self.sheet_name = self.config.get("sheet_name")  # If None, process first sheet or all
        self.header_row = int(self.config.get("header_row", 1)) - 1

    def connect(self) -> bool:
        if not os.path.exists(self.file_path):
            raise ConnectorError(self.connector_type, f"Excel file not found at: {self.file_path}")
        self._is_connected = True
        return True

    def disconnect(self) -> None:
        self._is_connected = False

    def test_connection(self) -> ConnectionTestResult:
        start = time.time()
        try:
            if not os.path.exists(self.file_path):
                return ConnectionTestResult(
                    success=False,
                    status="unhealthy",
                    latency_ms=(time.time() - start) * 1000,
                    message=f"File not found: {self.file_path}"
                )
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            latency = (time.time() - start) * 1000
            return ConnectionTestResult(
                success=True,
                status="healthy",
                latency_ms=round(latency, 2),
                message="Excel workbook verified accessible",
                details={"sheets": sheet_names, "file_path": self.file_path}
            )
        except Exception as exc:
            return ConnectionTestResult(
                success=False,
                status="unhealthy",
                latency_ms=(time.time() - start) * 1000,
                message=str(exc)
            )

    def discover_schema(self) -> SchemaDiscoveryResult:
        if not os.path.exists(self.file_path):
            self.connect()

        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        sheets_to_process = [self.sheet_name] if self.sheet_name else wb.sheetnames
        tables_meta: List[TableMeta] = []

        for sname in sheets_to_process:
            df_sample = pd.read_excel(self.file_path, sheet_name=sname, header=self.header_row, nrows=50)
            columns_meta: List[ColumnMeta] = []

            for col_name in df_sample.columns:
                series = df_sample[col_name].dropna()
                col_type = FieldType.STRING
                if not series.empty:
                    val = series.iloc[0]
                    if isinstance(val, bool):
                        col_type = FieldType.BOOLEAN
                    elif isinstance(val, int):
                        col_type = FieldType.INTEGER
                    elif isinstance(val, float):
                        col_type = FieldType.FLOAT
                    elif isinstance(val, (datetime, pd.Timestamp)):
                        col_type = FieldType.TIMESTAMP

                columns_meta.append(ColumnMeta(
                    name=str(col_name),
                    data_type=col_type,
                    is_nullable=df_sample[col_name].isnull().any(),
                    is_primary_key=(str(col_name).lower() in ("id", "key")),
                ))

            tables_meta.append(TableMeta(
                name=sname,
                schema_name="excel_sheet",
                columns=columns_meta,
                primary_keys=["id"] if any(c.name.lower() == "id" for c in columns_meta) else []
            ))

        wb.close()
        return SchemaDiscoveryResult(
            source_type=self.connector_type,
            tables=tables_meta,
            metadata={"file_path": self.file_path}
        )

    def preview_data(self, target: str, limit: int = 50) -> List[Dict[str, Any]]:
        sheet = target or self.sheet_name or 0
        df = pd.read_excel(self.file_path, sheet_name=sheet, header=self.header_row, nrows=limit)
        return df.where(pd.notnull(df), None).to_dict(orient="records")

    def extract_batch(
        self,
        target: str,
        chunk_size: int = 5000,
        incremental_column: Optional[str] = None,
        watermark_value: Optional[Any] = None,
    ) -> Generator[ExtractionChunk, None, None]:
        sheet = target or self.sheet_name or 0
        df_all = pd.read_excel(self.file_path, sheet_name=sheet, header=self.header_row)

        total_rows = len(df_all)
        chunk_idx = 0
        latest_watermark = watermark_value

        for start_idx in range(0, total_rows, chunk_size):
            df_chunk = df_all.iloc[start_idx : start_idx + chunk_size]
            records = df_chunk.where(pd.notnull(df_chunk), None).to_dict(orient="records")

            if incremental_column and records:
                latest_watermark = records[-1].get(incremental_column, latest_watermark)

            is_last = (start_idx + chunk_size) >= total_rows
            yield ExtractionChunk(
                chunk_index=chunk_idx,
                record_count=len(records),
                data=records,
                is_last_chunk=is_last,
                watermark_value=latest_watermark
            )
            chunk_idx += 1
